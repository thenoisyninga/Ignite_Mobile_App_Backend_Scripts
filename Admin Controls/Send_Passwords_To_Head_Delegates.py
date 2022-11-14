import smtplib
from email.message import EmailMessage

import openpyxl
import firebase_admin
from firebase_admin import firestore
from firebase_admin import credentials

cred = credentials.Certificate("serviceAccountKey.json")
firebase_admin.initialize_app(cred)
db = firestore.client()


# read data from the excel sheet, and sort it in a class

# Set up this function to send mails
def sendMail(delNum, email, password):
    # sendmail to 'email' with 'delNum' and 'password'
    notSendList = []
    failiureCount = 0
    successCount = 0
    sender = 'fpsalevelignite@fps.edu.pk'
    senderPassword = 'fpsignite@@2021'
    recipient_email = email
    msg = EmailMessage()
    msg['Subject'] = "Welcome to Ignite 8.0"
    msg['From'] = sender
    msg['To'] = recipient_email

    # Enter HTML message here....
    msg.set_content('''''', subtype='html')

    # Sending Messages...
    print(f'[+] Sending Message to {recipient_email} from delegation {delNum}...')
    with smtplib.SMTP("smtp.gmail.com", 587) as smtp:
        # with smtplib.SMTP("localhost", 1025) as smtp:
        try:
            successCount += 1
            smtp.starttls()
            smtp.login(sender, senderPassword)
            print(f"\nSending Message to {recipient_email}(Del {delNum}) with pass {password} ")
            # smtp.send_message(msg)

            if recipient_email == "sarim.ahmed19621@fpseagles.com":
                check = input("All set?")
        except:
            notSendList.append(recipient_email)
            print(f"\n[-] Message was not sent to {recipient_email}.\n")
            failiureCount += 1
            successCount -= 1


headPassDict = {}


class Delegate:
    def __init__(self, delNum, email, ):
        self.delNum = delNum
        self.email = email


# Registering Delegations...
for delNum in range(1, 51):
    if delNum == 1:
        delNum = "01"
    elif delNum == 2:
        delNum = "02"
    elif delNum == 3:
        delNum = "03"
    elif delNum == 4:
        delNum = "04"
    elif delNum == 5:
        delNum = "05"
    elif delNum == 6:
        delNum = "06"
    elif delNum == 7:
        delNum = "07"
    elif delNum == 8:
        delNum = "08"
    elif delNum == 9:
        delNum = "09"
    else:
        delNum = str(delNum)

    doc = db.collection('delegations_list').where('Delegation_Number', "==", delNum).get()
    doc = doc[0].to_dict()
    delNum = int(doc['Delegation_Number'])
    headDelPass = doc['Head_Del_Password']
    print(delNum, headDelPass)
    headPassDict[delNum] = headDelPass

print(headPassDict)

# delegationList = []
#
# path = "INDUSTRIES DATA.xlsx"
#
# wb_obj = openpyxl.load_workbook(path)
#
# delegation_emails = {}
#
# sheet_obj = wb_obj.active
#
# for i in range(50):
#     num = sheet_obj.cell(row=i, column=2).value
#     email = sheet_obj.cell(row=i, column=4).value
#     delegationList.append(Delegate(delNum=num, email=email))
#
#
# for delegate in delegationList:
#     sendMail(delNum=delegate.delNum, email=delegate.email, password=delPassDict[int(delegate.delNum)])
#
