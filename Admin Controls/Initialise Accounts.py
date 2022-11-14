import datetime
import firebase_admin
import openpyxl
from firebase_admin import firestore
from firebase_admin import credentials

print("[-] Setting things up...")

cred = credentials.Certificate("serviceAccountKey.json")
firebase_admin.initialize_app(cred)
db = firestore.client()


path = "INDUSTRIES DATA.xlsx"

wb_obj = openpyxl.load_workbook(path)

delegation_emails = {}

sheet_obj = wb_obj.active

kharchaDict = {}


def AddTransaction(ID, AmountChange, TransactionInfo):
    firestoreData = {
        "Amount": AmountChange,
        "Delegation_ID": ID,
        "Description": TransactionInfo,
        "transaction_time": datetime.datetime.now()
    }

    ICs = db.collection('delegations_ics_balance').where("Delegation_Number", "==", ID).get()
    doc = ICs[0].id
    ICs = ICs[0].to_dict()["Account_Balance"]

    ICs = ICs + int(AmountChange)

    data = {
        'Account_Balance': ICs,
    }
    db.collection('delegations_ics_balance').document(doc).update(data)

    db.collection('ic_transactions').document().set(firestoreData)

for delNum in range(1, 53):
    if delNum == 1:
        delNum = "01"
    elif delNum == 2:
        delNum = "02"
    elif delNum == 3:
        delNum = "03"
    elif delNum == 4:
        delNum = "04"
    elif delNum == 5:
        delNum = "05"
    elif delNum == 6:
        delNum = "06"
    elif delNum == 7:
        delNum = "07"
    elif delNum == 8:
        delNum = "08"
    elif delNum == 9:
        delNum = "09"
    else:
        delNum = str(delNum)

    data = {
        'Account_Balance': 0,
    }
    doc = db.collection('delegations_ics_balance').where('Delegation_Number', "==", delNum).get()
    id = doc[0].id
    db.collection('delegations_ics_balance').document(id).update(data)
    print(f"[+] Set DEL{delNum} ICs to 0.")


for i in range(50):
    row = (4*i)+2
    delNum = str(sheet_obj.cell(row=row, column=2).value)
    if int(delNum) == 1:
        delNum = "01"
    if int(delNum) == 2:
        delNum = "02"
    if int(delNum) == 3:
        delNum = "03"
    if int(delNum) == 4:
        delNum = "04"
    if int(delNum) == 5:
        delNum = "05"
    if int(delNum) == 6:
        delNum = "06"
    if int(delNum) == 7:
        delNum = "07"
    if int(delNum) == 8:
        delNum = "08"
    if int(delNum) == 9:
        delNum = "09"

    auctionKharcha = int(sheet_obj.cell(row=row, column=7).value)
    kharchaDict[delNum] = auctionKharcha

for i in range(1, 53):
    delNum = i
    if int(delNum) == 1:
        delNum = "01"
    if int(delNum) == 2:
        delNum = "02"
    if int(delNum) == 3:
        delNum = "03"
    if int(delNum) == 4:
        delNum = "04"
    if int(delNum) == 5:
        delNum = "05"
    if int(delNum) == 6:
        delNum = "06"
    if int(delNum) == 7:
        delNum = "07"
    if int(delNum) == 8:
        delNum = "08"
    if int(delNum) == 9:
        delNum = "09"
    AddTransaction(ID=str(delNum), AmountChange=1000, TransactionInfo="Initial ICs")
    print(f"\n[+] Gave DEL{delNum} 1000ICs")
    AddTransaction(ID=str(delNum), AmountChange=-int(kharchaDict[delNum]), TransactionInfo="Spent in Auction")
    print(f"[+] Took {kharchaDict[delNum]} form DEL{delNum}.")

