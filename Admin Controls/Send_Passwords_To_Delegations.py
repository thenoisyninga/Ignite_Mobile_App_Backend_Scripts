import smtplib
from email.message import EmailMessage

import openpyxl
import firebase_admin
from firebase_admin import firestore
from firebase_admin import credentials

cred = credentials.Certificate("serviceAccountKey.json")
firebase_admin.initialize_app(cred)

db = firestore.client()


# read data from the excel sheet, and sort it in a class

# Set up this function to send mails
def sendMail(delNum, email, password):
    # sendmail to 'email' with 'delNum' and 'password'
    notSendList = []
    failiureCount = 0
    successCount = 0
    sender = 'fpsalevelignite@fps.edu.pk'
    senderPassword = 'fpsignite@@2021'
    recipient_email = email
    msg = EmailMessage()
    msg['Subject'] = "Ignite App Head Delegate Credentials"
    msg['From'] = sender
    msg['To'] = recipient_email

    # Enter HTML message here....
    msg.set_content(f'''<!doctype html>
<!--Quite a few clients strip your Doctype out, and some even apply their own. Many clients do honor your doctype and it can make things much easier if you can validate constantly against a Doctype.-->
<html>
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>FPS Ignite</title>
<link rel="icon" type="image/png" href="favicon.ico">

<!-- Please use an inliner tool to convert all CSS to inline as inpage or external CSS is removed by email clients -->
<!-- important in CSS is used to prevent the styles of currently inline CSS from overriding the ones mentioned in media queries when corresponding screen sizes are encountered -->

<style type="text/css">
</style>
</head>
<body style="margin:0; padding:0;" bgcolor="#ffffff">
    <table style="min-width:320px;" width="100%" cellspacing="0" cellpadding="0" bgcolor="#ffffff">
      <!-- fix for gmail -->
      <tr>
        <td style="line-height:0;"><div style="display:none; white-space:nowrap; font:15px/1px courier;">&nbsp; &nbsp; &nbsp; &nbsp; &nbsp; &nbsp; &nbsp; &nbsp; &nbsp; &nbsp; &nbsp; &nbsp; &nbsp; &nbsp; &nbsp; &nbsp; &nbsp; &nbsp; &nbsp; &nbsp; &nbsp; &nbsp; &nbsp; &nbsp; &nbsp; &nbsp; &nbsp; &nbsp; &nbsp; &nbsp; &nbsp; &nbsp; &nbsp;</div></td>
      </tr>
      <!-- pre-header -->
      <!-- <tr>
        <td bgcolor="#000000">
          <table class="flexible" width="700" align="center" style="margin:0 auto;" cellpadding="0" cellspacing="0">
            <tr>
              <td align="center"  style="padding:9px 10px 11px; font:14px/19px Arial, Helvetica, sans-serif; color:#df7d26;">Not coming through? Click&nbsp;here&nbsp;to&nbsp;<a href="#" target="_blank" style="text-decoration:none; color:#fff;">view&nbsp;in&nbsp;browser</a></td>
            </tr>
          </table>
        </td>
      </tr>
      <tr> -->
        <td bgcolor="#fafafa" class="wrapper" style="padding:0 10px;">
          <table class="flexible" width="700" align="center" style="margin:0 auto;" cellpadding="0" cellspacing="0">
            <!-- fix for gmail -->
            <tr>
              <td class="hide">
                <table width="700" cellpadding="0" cellspacing="0" style="width:700px !important;">
                  <tr>
                    <td style="min-width:700px; font-size:0; line-height:0;">&nbsp;</td>
                  </tr>
                </table>
              </td>
            </tr>
            <!-- header -->
            <tr>
              <td class="header" style="padding:15px 0 15px;">
                
              </td>
            </tr>
            <!-- banner -->
            <tr>
              <td class="img-flex"><a href="http://fpsignite.com/" target="_blank"><img src="https://i.ibb.co/c2Ggj13/email-header.png" height="318" width="700" border="0" style="vertical-align:top;" alt="FPS IGNITE" /></a></td>

            </tr>
            <!-- section - 01 -->
            <tr>
              <td style="padding:0 0 50px;">
                <table width="100%" cellpadding="0" cellspacing="0">
                  <tr>
                    <td class="frame" bgcolor="#ffffff" style="padding:28px 40px 40px; border-radius:0 0 3px 3px;">
                      <table width="100%" cellpadding="0" cellspacing="0">
                        <tr>
                          <td align="center" style="padding:0 0 15px; font:bold 24px/26px Arial, Helvetica, sans-serif; color:#000;">Ignite App Head Delegate's Credentials &#57354;</td>
                        </tr>
                        <tr>
                          <td calign="center"  style="padding:0 0 43px; font:14px/25px Arial, Helvetica, sans-serif; color:#000000;">

<br>
Dear Delegates, <br><br>

The following are your Head Delegate's login credentials for the Ignite App<br><br>

Username: {delNum}<br>
Password: {password}<br><br>

<b>Important*: These credentials are to be kept highly confidential especially because this account has elevated priviliges. Each team must be logged in with this account in order to bookk their TVC Slots.</b><br>
<b> </b><br>

<!-- All the relevant information regarding the Auction is provided <a href="https://docs.google.com/document/d/17FUz7JAlPOM6yR35CmXGbjerwrmPY7b76Abrt3idCh8/edit?usp=sharing">here.</a><br><br> -->

Click here to download the app: <a href="https://drive.google.com/drive/folders/1hm1895aIwSnEcSTJluQkYEkuT857uH0n?usp=sharing">Ignite App</a><br><br>
Click here to watch a tutorial on how to install the app: <a href="https://youtu.be/UGymuTDhcTk">Tutorial</a><br><br>

Regards,<br>
FPS Ignite Registrations Department.
            
      <!-- footer -->
      <tr>
        <td bgcolor="#181a26">
          <table class="flexible" width="700" align="center" style="margin:0 auto;" cellpadding="0" cellspacing="0">
            <tr>
              <td class="footer" style="padding:50px 20px;">
                <table width="100%" cellpadding="0" cellspacing="0">
                  
                  <tr>
                    <td align="center" style="padding:0 0 15px; font:bold 16px/18px Arial, Helvetica, sans-serif; color:#fff;">INVEST | INSTILL | ILLUMINATE</td>
                    </tr>
                    <td align="center" style="font:14px/25px Arial, Helvetica, sans-serif; color:#fff;">
                      If you have any queries, contact us on <b><a href ="mailto:fpsalevelignite@fps.edu.pk">fpsalevelignite@fps.edu.pk</a><b> <br />
                      Or Call us on <br />
                      <a href="tel:03314464839">0336 IGNITE9</a><br/>
                      <a href="tel:03362250417">0336 2250417</a><br/>
                      <a href="tel:03353019641">0335 3019641</a><br/>
                      
                      Copyright &copy; 2022. All rights reserved.
                    </td>
                  </tr>
                </table>
              </td>
            </tr>
          </table>
        </td>
      </tr>
    </table>
  </body>
</html>
''', subtype='html')

    # Sending Messages...
    print(f'[+] Sending Message to {recipient_email} from delegation {delNum}...')
    with smtplib.SMTP("smtp.gmail.com", 587) as smtp:
        # with smtplib.SMTP("localhost", 1025) as smtp:
        try:
            successCount += 1
            smtp.starttls()
            smtp.login(sender, senderPassword)
            print(f"\nSending Message to {recipient_email}(Del {delNum}) with pass {password} ")
            smtp.send_message(msg)

            if recipient_email == "sarim.ahmed19621@fpseagles.com":
                check = input("All set?")
        except:
            notSendList.append(recipient_email)
            print(f"\n[-] Message was not sent to {recipient_email}.\n")
            failiureCount += 1
            successCount -= 1

class Delegate:
    def __init__(self, delNum, email, ):
        self.delNum = delNum
        self.email = email


delegationList = []

path = "INDUSTRIES DATA.xlsx"

wb_obj = openpyxl.load_workbook(path)

delegation_emails = {}

sheet_obj = wb_obj.active

for i in range(50):
    roww = (i*4)+2
    num = sheet_obj.cell(row=roww, column=2).value
    email = sheet_obj.cell(row=roww, column=4).value
    delegationList.append(Delegate(delNum=num, email=email))

delPassDict = {1: 'yYJxXhpzIn', 2: 'DAyUwLOqyy', 3: 'VxBjqOJTwg', 4: 'AJAypnAAbw', 5: 'KPprRcEumP', 6: 'zlVEivtAXg', 7: 'WtkbRfDZHG', 8: 'bfaIErZZNI', 9: 'TfUlZtczqm', 10: 'SLFVIsIpQg', 11: 'ATKUTgsHFc', 12: 'bFKWcjSFGv', 13: 'pNzLdqxHHO', 14: 'IwDVyzatrq', 15: 'nQpSQniTmI', 16: 'rwnhcyCQTb', 17: 'iiXpFIwToY', 18: 'acRaVMtayK', 19: 'gTjWnfOltG', 20: 'rrtqHieYxm', 21: 'LKKFPsfszL', 22: 'DvJQWMGwDj', 23: 'WrFISHbRvX', 24: 'exSOOrUAqs', 25: 'iSOUVRnemZ', 26: 'QayEntpnMC', 27: 'mQLYNsOJAk', 28: 'BCDBSrhCzX', 29: 'YGnOWoZMsn', 30: 'xAvIiEiUdj', 31: 'JGXiAMqbnl', 32: 'lpNvpjQUIY', 33: 'reWUYkaLIr', 34: 'vLBfXXUgOj', 35: 'KpohpsUjCO', 36: 'wCLmMwHcIS', 37: 'IOMKdlKkFE', 38: 'OnNTVVBcGs', 39: 'aVyRAKnLJK', 40: 'wIRxFkdrWz', 41: 'lgRFoJniSb', 42: 'pBNTtzHUMn', 43: 'nUctuindiS', 44: 'fBCsWXjWKh', 45: 'kIKFVNjXnt', 46: 'DNDBriHuEl', 47: 'WERCdrcHrr', 48: 'TITrrBSlQE', 49: 'ocPdkwJfsf', 50: 'CQfmgqvlyn'}


sendMail(delNum=51, email='sarim.ahmed19621@fpseagles.com', password='1234')
sendMail(delNum=51, email='ebrahim.baig@fps.edu.pk', password='1234')

x = input("(Y/n): ")

if x == 'Y':
    for delegate in delegationList:
        # print(delegate.delNum, delegate.email, delPassDict[delegate.delNum])
        sendMail(delNum=delegate.delNum, email=delegate.email, password=delPassDict[int(delegate.delNum)])


